import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
import re
from io import BytesIO

st.title("Roll Number Filter Tool")

# Use a form to group inputs and a submit button
with st.form(key='roll_number_form'):
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    target = st.number_input("Enter roll number (1-150)", min_value=1, max_value=150, step=1)
    submit_button = st.form_submit_button(label='Submit')

if submit_button and uploaded_file:
    # Load workbook from the uploaded file (from BytesIO)
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    # Define the highlight fill color (yellow).
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Process rows starting from row 3 (assuming rows 1 and 2 are headers)
    for row in ws.iter_rows(min_row=3, min_col=4, max_col=ws.max_column):
        for cell in row:
            original_value = cell.value
            if original_value is None:
                continue

            s = str(original_value).strip()
            # Use regex to extract tokens that are numbers or ranges (e.g., "6" or "1-10")
            tokens = re.findall(r'\d+(?:-\d+)?', s)
            has_target = False
            for token in tokens:
                if '-' in token:
                    try:
                        low_str, high_str = token.split('-')
                        low = int(low_str.strip())
                        high = int(high_str.strip())
                        # If the range includes the target roll
                        if low <= target <= high:
                            has_target = True
                            break
                    except Exception:
                        continue
                else:
                    try:
                        # Only consider if the token is exactly equal to the target.
                        if int(token.strip()) == target:
                            has_target = True
                            break
                    except Exception:
                        continue

            if has_target:
                # Set the cell value to the target and highlight it.
                cell.value = str(target)
                cell.fill = highlight_fill
            else:
                # Clear cells that don't include the target roll.
                cell.value = None

    # Save the modified workbook to a BytesIO stream.
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        label="Download Filtered Excel",
        data=output,
        file_name="Filtered_Internship_Postings_Colored.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
